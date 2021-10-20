import requests
import lxml.html
import random, sys
from time import sleep, strftime, time
from openpyxl import Workbook

output_file_name = "data_{}.xlsx".format(strftime("%Y%m%d%H%M%S"))
help_text = """
MISSING config.py.
Create config.py with following content:

base_url="https://example.com"
target_url="{0}/path"
data_headers = ["column_name_1", "column_name_2"]
"""
try:
    import config
except:
    sys.exit(help_text)

headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9,bn-BD;q=0.8,bn;q=0.7",
    "Upgrade-Insecure-Requests": "1",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.81 Safari/537.36",
    'referer': "https://www.google.com/"
}

base_url = config.base_url


def wait():
    sleep(random.uniform(1, 5))


def main():
    script_start_rime = time()
    print("WORKING", end="")
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data.append(config.data_headers)
    ws_error = wb.create_sheet("Error")
    ws_error.append(["Parent", "Target", "Response"])
    ws_meta = wb.create_sheet("Meta")
    web_hit = 0

    for page in range(1, 78):
        print(".", end="")
        wait()
        response = requests.get(base_url, headers=headers, params={"page": page})
        web_hit += 1
        parent_page_url = response.url

        if response.status_code != 200:
            ws_error.append([parent_page_url, parent_page_url, response.status_code])
            continue

        html_content = lxml.html.fromstring(response.content)
        root_div = html_content.xpath("/html/body/main/div/section/div/div[2]")[0]
        for div in root_div:
            for a in div:
                target_url = config.target_url.format(a.get("href"))
                print(".", end="")
                wait()
                response = requests.get(target_url, headers=headers, params={})
                web_hit += 1
                if response.status_code != 200:
                    ws_error.append([parent_page_url, target_url, response.status_code])
                    continue

                root = lxml.html.fromstring(response.content)
                table_tbody = root.xpath("/html/body/main/div/section/div[2]/div/div[2]/table/tbody")[0]

                for table_tr in table_tbody:
                    table_row = []
                    for table_td in table_tr:
                        cell_data = table_td.text_content().strip().replace("\n", "").replace("  ", "")
                        table_row.append(cell_data)
                    ws_data.append(table_row)
                    print(".", end="")

    script_end_time = time()
    ws_meta.append(["Date", "Start", "End", "Runtime (sec)", "Web Hit"])
    ws_meta.append(
        [strftime("%Y-%m-%d"), script_start_rime, script_end_time, script_end_time - script_start_rime, web_hit])
    wb.save(output_file_name)
    print("DONE")


if __name__ == "__main__":
    main()
